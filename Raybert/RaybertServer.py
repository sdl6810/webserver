from flask import Flask
from flask import render_template
from flask import session, request
import time
from polynomialEncryption import *
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key="random nonsense"

@app.route('/my-form/')
def my_form():
    return render_template('my-form.html',session=session)

@app.route('/my-form/',methods=['POST'])
def my_form_post():
    text = request.form['text']
    processed_text = text.upper()
    return processed_text

@app.route('/submit', methods=['POST'])
def submit():
    wb = load_workbook('/home/sdl5384/Desktop/Python_SRC/webserver/Raybert/messageData.xlsx')
    ws = wb.active
    text = format(request.form['text'])
    e = polynomialEncryption(text)
    text = e.encryptText()
    nextRow = ws.max_row+1
    ws['A'+str(nextRow)].value = 'Timestamp'
    ws['B'+str(nextRow)].value = e.newMsg

    wb.save('/home/sdl5384/Desktop/Python_SRC/webserver/Raybert/messageData.xlsx')
    return render_template('/after-data-entered.html',session=session)

# @app.route('/rollTheDice/')
# def rollTheDice():
#     # msg = "The numbers you rolled are the following: "
#     # die = []
#     # for i in range(0,15):
#     #     die.append(r.randint(1,6))
#     #     msg = msg + "2 " + session["die3"] + session["die4"] + session["die5"] + str(die[i])

#     if "count" not in session:
#         session["count"] = 0
#     else:
#         session["count"] = session["count"] + 1

#     if "diceAvg" not in session:
#         session["diceAvg"] = 0

#     if "die1" not in session or "die2" not in session or "die3" not in session or "die4" not in session or "die5" not in session:
#         session["die1"] = 0
#         session["die2"] = 0
#         session["die3"] = 0
#         session["die4"] = 0
#         session["die5"] = 0
#     else:
#         session["die1"] = r.randint(1,6)
#         session["die2"] = r.randint(1,6)
#         session["die3"] = r.randint(1,6)
#         session["die4"] = r.randint(1,6)
#         session["die5"] = r.randint(1,6)

#         session["diceAvg"] = (session["die1"] + session["die2"] + session["die3"] + session["die4"] + session["die5"])/5
        
#     return render_template("diceSimulation.html",session=session)
app.run(host='0.0.0.0', port=5000)