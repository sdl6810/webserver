from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import random
import math

class polynomialEncryption():
	def __init__(self,message):
		self._message = message
		self.lowercaseLetters = ['a','b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
		self.capitals = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
		self.numbers = ['0','1','2','3','4','5','6','7','8','9']
		self.punctuation = ["'", "?", "!", "*", ".", "-", "_", "(", ")", "$", "#", "&", "-", "+", "/"]
		self.degree = random.randint(1,10)
		self.leadingCoeff = random.randint(1,10)
		self.coefficients = [self.leadingCoeff]
		self.generateCoefficients()
		self.exponents = [self.degree]
		self.generateExponents()
		self.newMsg = ''
		self.inputValueChosen = random.randint(1,100)

	def generateExponents(self):
		i = 2*self.leadingCoeff
		while (i > 0):
			self.exponents.append(random.randint(1,i))
			i = i - 1

	def generateCoefficients(self):
		j = 2*self.leadingCoeff
		#self.coefficients.append(self.leadingCoeff)
		while (j > 0):
			self.coefficients.append(random.randint(1,j))
			j = j - 1

	def calculatePermutation(self,value,divisor):
		result = 0.0
		for i in range(0,len(self.exponents)-1):
			result = (result + self.coefficients[i]*value+math.pow(value,self.exponents[i]))%divisor

		return result

	def calculatePermutationRevised(self):
		result = 0.0
		for i in range(0,len(self.exponents)-1):
			result = result + self.coefficients[i]*self.inputValueChosen+math.pow(self.inputValueChosen,self.exponents[i])

		return result

	def secondLevelPermutation(self, originalLocation, firstLevelValue):
		difference = originalLocation - firstLevelValue
		result = (difference - 5)*(difference + 12)*(difference - 13)

		return int(result)

	def searchFor(self, array, character):
		characterIsHere = 0
		for q in range(0, len(array)-1):
			if character == array[q]:
				characterIsHere = q
		return characterIsHere

	def encryptText(self):
		newCapitalLetter = self.calculatePermutationRevised()%len(self.capitals)
		newLowerCaseLetter = self.calculatePermutationRevised()%len(self.lowercaseLetters)
		newNumber = self.calculatePermutationRevised()%len(self.numbers)
		newPunctuation = self.calculatePermutationRevised()%len(self.punctuation)

		originalLocation = 0

		for i in range(0, len(self._message)):
			if self._message[i] == " ":
				self.newMsg = self.newMsg + '0'
			elif self._message[i] in self.numbers:
				originalLocation = self.searchFor(self.numbers,self._message[i])
				self.newMsg = self.newMsg + self.numbers[self.secondLevelPermutation(originalLocation,int(newNumber))%len(self.numbers)]
			elif self._message[i] in self.lowercaseLetters:
				originalLocation = self.searchFor(self.lowercaseLetters,self._message[i])
				self.newMsg = self.newMsg + self.lowercaseLetters[self.secondLevelPermutation(originalLocation,int(newLowerCaseLetter))%len(self.lowercaseLetters)]
			elif self._message[i] in self.capitals:
				originalLocation = self.searchFor(self.capitals,self._message[i])
				self.newMsg = self.newMsg + self.capitals[self.secondLevelPermutation(originalLocation,int(newCapitalLetter))%len(self.capitals)]
			elif self._message[i] in self.punctuation:
				originalLocation = self.searchFor(self.punctuation,self._message[i])
				self.newMsg = self.newMsg + self.punctuation[self.secondLevelPermutation(originalLocation,int(newPunctuation))%len(self.punctuation)]

	def exportMsgToXl(self, filepath):
		encryptedData = Workbook()
		text = encryptedData.active

		row = 1
		for i in range(0, len(self.newMsg)-1):
			if i%13 == 0:
				row = row + 1
				text[self.capitals[i%13]+str(row)] = self.newMsg[i]
			else:
				text[self.capitals[i%13]+str(row)] = self.newMsg[i]

		#save as .xlsx
		encryptedData.save(filepath+'/experiment.xlsx')
		#save as another Macro-enabled file (.xlsm)
		encryptedData.save(filepath+'/experiment.xlsm')

		lastRow = text.max_row
		# #include encryption components two rows down in spreadsheet
		text[get_column_letter(2)+str(lastRow+2)] = 'exponents'
		text[get_column_letter(2)+str(lastRow+3)] = 'coefficients'
		for x in range(0, len(self.exponents)-1):
			text[get_column_letter(x+3)+str(lastRow+2)] = self.exponents[x]
			text[get_column_letter(x+3)+str(lastRow+3)] = self.coefficients[x]
		
		#save again as .xlsx
		encryptedData.save(filepath+'/experiment.xlsx')
		#save again as another Macro-enabled file (.xlsm)
		encryptedData.save(filepath+'/experiment.xlsm')

# def main():
# 	poly = polynomialEncryption('This is a test of the emergency broadcast system.')
# 	poly.encryptText()
# 	print(poly.newMsg)
# 	poly.exportMsgToXl('/home/sdl5384/Desktop')

# main()